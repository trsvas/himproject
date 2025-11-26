import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
import sqlite3
import math
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image
import os
import time
import tracemalloc
import psutil


class MyApp:
    def __init__(self, root):
        # Установка заголовка окна
        self.root = root
        self.root.title("Данные модели")
        self.root.geometry("800x300")

        self._center_window(self.root)

        # Создание Combobox в верхней части окна
        self.combo_label = tk.Label(root, text="Выберите материал:")
        self.combo_label.pack(anchor="nw", padx=10, pady=5)

        self.combobox = ttk.Combobox(root)
        self.combobox.pack(anchor="nw", padx=10)
        self.combobox.bind("<<ComboboxSelected>>", self.show_data)

        self.load_materials()  # Загрузка материалов при инициализации

        # Создание первого фрейма для вводимых параметров
        self.frame1 = tk.Frame(root, width=350, height=250, relief=tk.GROOVE, borderwidth=2)
        self.frame1.pack_propagate(False)
        self.frame1.pack(side=tk.LEFT, padx=10, pady=10)

        self.frame1_label = tk.Label(self.frame1, text="Вводимые параметры")
        self.frame1_label.pack(pady=5)

        self.temp_frame = tk.Frame(self.frame1)
        self.temp_frame.pack(side=tk.LEFT, padx=10)

        # Столбец для температур
        self.label_Tmin = tk.Label(self.temp_frame, text="Tmin, 'С")
        self.label_Tmin.pack(side=tk.TOP)

        self.entry_Tmin = tk.Entry(self.temp_frame)
        self.entry_Tmin.pack(side=tk.TOP, padx=5)

        self.label_Tst = tk.Label(self.temp_frame, text="Tst, 'С")
        self.label_Tst.pack(side=tk.TOP)

        self.entry_Tst = tk.Entry(self.temp_frame)
        self.entry_Tst.pack(side=tk.TOP, padx=5)

        self.label_Tmax = tk.Label(self.temp_frame, text="Tmax, 'С")
        self.label_Tmax.pack(side=tk.TOP)

        self.entry_Tmax = tk.Entry(self.temp_frame)
        self.entry_Tmax.pack(side=tk.TOP, padx=5)

        # Фрейм для расхода (Q)
        self.q_frame = tk.Frame(self.frame1)
        self.q_frame.pack(side=tk.LEFT, padx=10)

        self.label_qmin = tk.Label(self.q_frame, text="Qmin, л/мин")
        self.label_qmin.pack(side=tk.TOP)

        self.entry_qmin = tk.Entry(self.q_frame)
        self.entry_qmin.pack(side=tk.TOP, padx=5)

        self.label_qst = tk.Label(self.q_frame, text="Qst, л/мин")
        self.label_qst.pack(side=tk.TOP)

        self.entry_qst = tk.Entry(self.q_frame)
        self.entry_qst.pack(side=tk.TOP, padx=5)

        self.label_qmax = tk.Label(self.q_frame, text="Qmax, л/мин")
        self.label_qmax.pack(side=tk.TOP)

        self.entry_qmax = tk.Entry(self.q_frame)
        self.entry_qmax.pack(side=tk.TOP, padx=5)

        # Создание второго фрейма для параметров по умолчанию
        self.frame2 = tk.Frame(root, width=250, height=300, relief=tk.GROOVE, borderwidth=2)
        self.frame2.pack_propagate(False)
        self.frame2.pack(side=tk.LEFT, padx=10, pady=10)

        self.frame2_label = tk.Label(self.frame2, text="Параметры по умолчанию")
        self.frame2_label.pack(pady=5)

        # Ввод 4 значений во второй фрейм
        self.label_td = tk.Label(self.frame2, text="Td,'С")
        self.label_td.pack(anchor='w')

        self.entry_td = tk.Entry(self.frame2)
        self.entry_td.pack()

        self.label_ed = tk.Label(self.frame2, text="Ed, Дж/моль")
        self.label_ed.pack(anchor='w')

        self.entry_ed = tk.Entry(self.frame2)
        self.entry_ed.pack()

        self.label_v = tk.Label(self.frame2, text="Ve, л")
        self.label_v.pack(anchor='w')

        self.entry_v = tk.Entry(self.frame2)
        self.entry_v.pack()

        self.label_t = tk.Label(self.frame2, text="t, мин")
        self.label_t.pack(anchor='w')

        self.entry_t = tk.Entry(self.frame2)
        self.entry_t.pack()

        # Кнопка расчета
        self.calculate_button = tk.Button(root, text="Расчет", command=self.calculate)
        self.calculate_button.pack(side=tk.BOTTOM, pady=10)

    def load_materials(self):
        # Создаем соединение с базой данных
        conn = sqlite3.connect('info.db')
        cursor = conn.cursor()

        # Получаем все названия материалов из столбца name
        cursor.execute("SELECT name FROM coefficients")
        materials = cursor.fetchall()

        # Добавляем названия в Combobox
        self.combobox['values'] = [material[0] for material in materials]

        conn.close()

    def calculate(self):
        global Tmin, Tmax, Tst, qmax, qmin, qst, t, Ed, Td, V
        try:
            # Получаем значения из полей Entry
            qmin = float(self.entry_qmin.get())
            qmax = float(self.entry_qmax.get())
            qst = float(self.entry_qst.get())
            Tmin = float(self.entry_Tmin.get())
            Tmax = float(self.entry_Tmax.get())
            Tst = float(self.entry_Tst.get())
            # Получение других параметров из полей
            Ed = float(self.entry_ed.get())
            t = float(self.entry_t.get())
            Td = float(self.entry_td.get())
            V = float(self.entry_v.get())

            # Проверка на положительные значения
            if all(val > 0 for val in [qmin, qmax, qst, Tmin, Tmax, Tst, Ed, t, Td, V]):
                self.show_result_window()

            else:
                messagebox.showwarning("Ошибка ввода", "Все значения должны быть положительными.")

        except ValueError:
            messagebox.showwarning("Ошибка ввода", "Проверьте вводимые значения")

    def show_result_window(self):
        # Создание нового окна с настройками модальности
        result_window = tk.Toplevel(self.root)
        result_window.title("Результат расчета")
        result_window.minsize(width=400, height=250)

        # Делаем окно модальным и поверх главного
        result_window.transient(self.root)
        result_window.grab_set()
        result_window.focus_set()

        self._center_window(result_window)

        # Создание кнопок
        report_button = tk.Button(result_window, text="Таблица", command=self.show_otch_window)
        report_button.pack(pady=10)

        report_button = tk.Button(result_window, text="Графики", command=self.plot_graph)
        report_button.pack(pady=10)

        report_button = tk.Button(result_window, text="Отчет в Excel", command=self.save_ex_window)
        report_button.pack(pady=10)

        report_button = tk.Button(result_window, text="Оценка экономичности", command=self.measure_performance)
        report_button.pack(pady=10)

        exit_button = tk.Button(result_window, text="Выйти из программы",
                                command=self.exit_program, bg="lightcoral")
        exit_button.pack(pady=10)

        # Обработчик закрытия окна
        result_window.protocol("WM_DELETE_WINDOW", lambda: self._on_child_close(result_window))

    def _on_child_close(self, window):
        """Обработчик закрытия дочернего окна"""
        window.grab_release()
        window.destroy()
        self.root.focus_set()

    def exit_program(self):
        """Выход из программы"""
        if messagebox.askokcancel("Выход", "Вы уверены, что хотите выйти?"):
            self.root.destroy()

    def plot_graph(self):
        q_values = [qmin, (qmin + qmax) / 2, qmax]
        t_values = np.arange(Tmin, Tmax, Tst)
        q_values_for_graph_2 = np.arange(qmin, qmax + 1, qst)
        t_values_for_graph_2 = [Tmin, (Tmin + Tmax) / 2, Tmax]

        # Создаем фигуру и 2 подграфика
        fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(10, 9))

        # График I от T
        for q in q_values:
            I_values = [
                (V / (t * q)) * np.exp(Ed / (8.31 * (T + 273.15) * (Td + 273.15)) * (T - Td)) * 100
                for T in t_values
            ]
            ax1.plot(t_values, I_values, label=f'Q = {q}')

        ax1.set_title("Зависимость I от T")
        ax1.set_xlabel("Температура,(°С)")
        ax1.set_ylabel("Индекс деструкции, %")
        ax1.legend()
        ax1.grid()

        # График I от q
        for T in t_values_for_graph_2:
            I_values = [
                (V / (t * q)) * np.exp(Ed / (8.31 * (T + 273.15) * (Td + 273.15)) * (T - Td)) * 100
                for q in q_values_for_graph_2
            ]
            ax2.plot(q_values_for_graph_2, I_values, label=f'T = {T}')

        ax2.set_title("Зависимость I от Q")
        ax2.set_xlabel("Расход потока материала, (л/мин)")
        ax2.set_ylabel("Индекс деструкции, %")
        ax2.legend()
        ax2.grid()

        plt.tight_layout()
        plt.subplots_adjust(hspace=0.2)
        plt.savefig("графики.png")
        plt.show()

    def _center_window(self, window):
        """Центрирование окна на экране"""
        window.update_idletasks()
        width = window.winfo_width()
        height = window.winfo_height()
        x = (window.winfo_screenwidth() // 2) - (width // 2)
        y = (window.winfo_screenheight() // 2) - (height // 2)
        window.geometry(f"{width}x{height}+{x}+{y}")

    def save_ex_window(self):
        # Создаем временное окно для сообщения
        temp_window = tk.Toplevel(self.root)
        temp_window.withdraw()  # Скрываем временное окно

        try:
            q_values = np.arange(qmin, qmax + 1, qst)
            t_values = np.arange(Tmin, Tmax + 1, Tst)
            material = str(self.combobox.get())

            # Создаем DataFrame для Excel
            data = []
            for q in q_values:
                row_data = []
                for T in t_values:
                    I = (V / (t * q)) * math.exp(Ed / (8.31 * (T + 273.15) * (Td + 273.15)) * (T - Td)) * 100
                    row_data.append(I)
                data.append(row_data)

            # Создание DataFrame
            df = pd.DataFrame(data, columns=t_values, index=q_values)

            # Сохранение в Excel
            excel_filename = f"отчет {material}.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "Данные"

            # Запись заголовков для первого столбца
            ws.cell(row=1, column=1, value="Q")
            for i, q_value in enumerate(q_values):
                ws.cell(row=i + 2, column=1, value=q_value)

            for j, t_value in enumerate(t_values):
                ws.cell(row=1, column=j + 2, value=t_value)

            # Запись DataFrame в Excel
            for r_idx, row in enumerate(df.iterrows()):
                for c_idx, value in enumerate(row[1]):
                    ws.cell(row=r_idx + 2, column=c_idx + 2, value=value)

            # Вставка изображения графика в Excel
            if os.path.exists("графики.png"):
                img = Image("графики.png")
                img.width = 400
                img.height = 600
                ws.add_image(img, "B25")

            # Сохранить файл
            wb.save(excel_filename)

            # Показываем сообщение поверх всех окон
            temp_window.destroy()
            messagebox.showinfo("Сообщение", f"Отчет '{excel_filename}' создан", parent=self.root)

        except Exception as e:
            temp_window.destroy()
            messagebox.showerror("Ошибка", f"Ошибка при создании отчета: {str(e)}", parent=self.root)

    def show_otch_window(self, *args):
        report_window = tk.Toplevel(self.root)
        report_window.title("Таблица")
        report_window.geometry("1000x500")

        # Настройки модальности
        report_window.transient(self.root)
        report_window.grab_set()
        report_window.focus_set()

        # Создаем таблицу
        tree = ttk.Treeview(report_window, height=15)
        tree.pack(expand=True, fill="both")

        # Создаем вертикальные скроллбары
        y_scroll = ttk.Scrollbar(report_window, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=y_scroll.set)
        y_scroll.pack(side='right', fill='y')

        # Получаем значения из полей ввода
        qmin = float(self.entry_qmin.get())
        qmax = float(self.entry_qmax.get())
        qst = float(self.entry_qst.get())
        Tmin = float(self.entry_Tmin.get())
        Tmax = float(self.entry_Tmax.get())
        Tst = float(self.entry_Tst.get())
        Ed = float(self.entry_ed.get())
        t = float(self.entry_t.get())
        Td = float(self.entry_td.get())
        V = float(self.entry_v.get())

        # Генерируем значения Q и T
        q_values = np.arange(qmin, qmax + 0.5, qst)
        t_values = np.arange(Tmin, Tmax + 1, Tst)

        # Заголовки таблицы
        columns = ['Q'] + [f'{T:.1f}' for T in t_values]
        tree['columns'] = columns

        # Настройка заголовков
        for col in columns:
            tree.heading(col, text=col)
            tree.column(col, anchor="center", width=50)

        # Расчет и добавление значений в таблицу
        for q in q_values:
            row_data = [f'{q:.1f}']
            for T in t_values:
                if t * q != 0:
                    I = (V / (t * q)) * math.exp(Ed / (8.31 * (T + 273.15) * (Td + 273.15)) * (T - Td)) * 100
                else:
                    I = 0.0
                row_data.append(f'{I:.2f}')
            tree.insert("", "end", values=row_data)

        # Удалить первый пустой столбец
        tree.column("#0", width=0, stretch=tk.NO)
        tree.column("Q", width=30, anchor="center")

        # Обработчик закрытия окна
        report_window.protocol("WM_DELETE_WINDOW", lambda: self._on_child_close(report_window))

    def show_data(self, event):
        global selected_material
        selected_material = self.combobox.get()

        conn = sqlite3.connect('info.db')
        cursor = conn.cursor()
        cursor.execute("SELECT temp, energy, time, V FROM coefficients WHERE name = ?", (selected_material,))
        result = cursor.fetchone()

        if result:
            temp, energy, time, V = result
            self.entry_td.delete(0, tk.END)
            self.entry_td.insert(0, temp)
            self.entry_ed.delete(0, tk.END)
            self.entry_ed.insert(0, energy)
            self.entry_t.delete(0, tk.END)
            self.entry_t.insert(0, time)
            self.entry_v.delete(0, tk.END)
            self.entry_v.insert(0, V)

        conn.close()

    def measure_performance(self):
        # Создаем временное окно для сообщения
        temp_window = tk.Toplevel(self.root)
        temp_window.withdraw()

        try:
            conn = sqlite3.connect('info.db')
            cur = conn.cursor()
            cur.execute("SELECT temp, energy, time, V FROM coefficients WHERE name = ?",
                        (selected_material,))
            materials = cur.fetchone()
            if not materials:
                messagebox.showerror("Ошибка", "Коэффициенты для данного типа материала не найдены.", parent=self.root)
                return

            # Начало отсчета времени и памяти для всего процесса
            total_start_time = time.time()
            tracemalloc.start()

            # 1. ВРЕМЯ ВЫЧИСЛЕНИЯ МОДЕЛИ
            computation_start_time = time.time()

            # Вычисляем количество точек
            t_points = len(np.arange(Tmin, Tmax + 1, Tst))
            q_points = len(np.arange(qmin, qmax + 1, qst))

            # Общее количество арифметических операций
            total_operations = t_points * q_points * 12

            # Выполняем вычисления модели
            t_values = np.arange(Tmin, Tmax + 1, Tst)
            q_values = np.arange(qmin, qmax + 1, qst)

            # Создаем массивы для хранения результатов
            results = []
            for q in q_values:
                row_results = []
                for T in t_values:
                    I = (V / (t * q)) * math.exp(Ed / (8.31 * (T + 273.15) * (Td + 273.15)) * (T - Td)) * 100
                    row_results.append(I)
                results.append(row_results)

            computation_end_time = time.time()
            computation_time = computation_end_time - computation_start_time

            # Создание графиков
            fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(10, 9))

            # График 1: I от T
            q_graph_values = [qmin, (qmin + qmax) / 2, qmax]
            for q in q_graph_values:
                I_values = [
                    (V / (t * q)) * np.exp(Ed / (8.31 * (T + 273.15) * (Td + 273.15)) * (T - Td)) * 100
                    for T in t_values
                ]
                ax1.plot(t_values, I_values, label=f'Q = {q}')

            ax1.set_title("Зависимость I от T")
            ax1.set_xlabel("Температура,(°С)")
            ax1.set_ylabel("Индекс деструкции, %")
            ax1.legend()
            ax1.grid()

            # График 2: I от Q
            q_values_for_graph_2 = np.arange(qmin, qmax + 1, qst)
            t_graph_values = [Tmin, (Tmin + Tmax) / 2, Tmax]
            for T in t_graph_values:
                I_values = [
                    (V / (t * q)) * np.exp(Ed / (8.31 * (T + 273.15) * (Td + 273.15)) * (T - Td)) * 100
                    for q in q_values_for_graph_2
                ]
                ax2.plot(q_values_for_graph_2, I_values, label=f'T = {T}')

            ax2.set_title("Зависимость I от Q")
            ax2.set_xlabel("Расход потока материала, (л/мин)")
            ax2.set_ylabel("Индекс деструкции, %")
            ax2.legend()
            ax2.grid()

            plt.tight_layout()
            plt.subplots_adjust(hspace=0.2)
            plt.savefig("графики_производительность.png")
            plt.close()

            # Создание DataFrame для таблицы
            df = pd.DataFrame(results, columns=t_values, index=q_values)
            table_data = []
            for q in q_values:
                row_data = [f'{q:.1f}']
                for T in t_values:
                    I = (V / (t * q)) * math.exp(Ed / (8.31 * (T + 273.15) * (Td + 273.15)) * (T - Td)) * 100
                    row_data.append(f'{I:.2f}')
                table_data.append(row_data)

            excel_filename = "отчет_производительность.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "Данные"

            # Запись данных в Excel
            ws.cell(row=1, column=1, value="Q")
            for i, q_value in enumerate(q_values):
                ws.cell(row=i + 2, column=1, value=q_value)

            for j, t_value in enumerate(t_values):
                ws.cell(row=1, column=j + 2, value=t_value)

            for r_idx, row in enumerate(results):
                for c_idx, value in enumerate(row):
                    ws.cell(row=r_idx + 2, column=c_idx + 2, value=value)

            wb.save(excel_filename)

            # Конец общего отсчета
            total_end_time = time.time()
            total_time = total_end_time - total_start_time

            # Получение данных об использовании памяти
            current, peak = tracemalloc.get_traced_memory()
            tracemalloc.stop()

            memory_usage_mb = peak / 1024 / 1024
            memory_usage_kb = peak / 1024

            # Формирование отчета о производительности
            performance_message = (
                f"• Общее время выполнения: {total_time:.4f} с\n"
                f"• Использовано памяти: {memory_usage_mb:.2f} МБ ({memory_usage_kb:.2f} КБ)\n"
                f"• Арифметических операций: {total_operations:,}\n"
            )

            # Очистка временных файлов
            if os.path.exists("графики_производительность.png"):
                os.remove("графики_производительность.png")
            if os.path.exists("отчет_производительность.xlsx"):
                os.remove("отчет_производительность.xlsx")

            temp_window.destroy()
            messagebox.showinfo("Оценка экономичности", performance_message, parent=self.root)

        except Exception as e:
            temp_window.destroy()
            messagebox.showerror("Ошибка", f"Ошибка при оценке производительности: {str(e)}", parent=self.root)

    def run(self):
        self.root.mainloop()


if __name__ == "__main__":
    root = tk.Tk()
    app = MyApp(root)
    app.run()